import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .models import Turma, Aluno

def exportar_excel_turma(request, turma_id):
    try:
        turma = Turma.objects.get(pk=turma_id)
    except Turma.DoesNotExist:
        return HttpResponse('Turma não encontrada.', status=404)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Turma {turma.turma}"

    # Cabeçalho da turma
    ws.append(["Série", "Turma", "Turno", "Professor", "Representante"])
    ws.append([
        turma.serie,
        turma.turma,
        turma.turno,
        turma.professor_principal.nome_professor if turma.professor_principal else '',
        turma.representante.nome_aluno if turma.representante else ''
    ])
    ws.append([])
    ws.append(["Alunos"])
    ws.append(["Nome", "Matrícula"])
    for aluno in turma.alunos.filter(ativo=True):
        ws.append([aluno.nome_aluno, aluno.matricula_aluno])

    # Ajusta largura das colunas
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=turma_{turma_id}.xlsx'
    wb.save(response)
    return response
